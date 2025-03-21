import pandas as pd
from datetime import datetime
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def leer_datos_campanas(ruta_archivo):
    """Lee los datos de campañas desde un archivo CSV."""
    try:
        # Leer el archivo CSV
        df = pd.read_csv(ruta_archivo)
        return df
    except Exception as e:
        print(f"Error al leer el archivo CSV: {str(e)}")
        raise

def procesar_datos(df):
    """Procesa los datos de las campañas."""
    try:
        # Convertir columnas de fecha a datetime
        if 'fecha_inicio' in df.columns:
            df['fecha_inicio'] = pd.to_datetime(df['fecha_inicio'])
        if 'fecha_fin' in df.columns:
            df['fecha_fin'] = pd.to_datetime(df['fecha_fin'])
        
        # Convertir columnas numéricas
        columnas_numericas = ['gasto', 'impresiones', 'clicks', 'alcance', 'cpc', 'ctr', 'frecuencia']
        for columna in columnas_numericas:
            if columna in df.columns:
                df[columna] = pd.to_numeric(df[columna], errors='coerce')
        
        return df
    except Exception as e:
        print(f"Error al procesar los datos: {str(e)}")
        raise

def verificar_rutas(nombre_archivo):
    """
    Verifica y crea las rutas necesarias para los archivos.
    
    Args:
        nombre_archivo (str): Nombre del archivo CSV a procesar
        
    Returns:
        str: Ruta completa del archivo si existe, None si no existe
    """
    # Crear ruta completa
    ruta_entrada = os.path.join('campanas', nombre_archivo)
    
    # Verificar si la carpeta existe, si no, crearla
    if not os.path.exists('campanas'):
        os.makedirs('campanas')
        print("Se ha creado la carpeta 'campanas'")
        print("Por favor, coloca tu archivo CSV en la carpeta 'campanas'")
        return None
    
    # Verificar si el archivo existe
    if not os.path.exists(ruta_entrada):
        print(f"Error: No se encontró el archivo '{nombre_archivo}' en la carpeta 'campanas'")
        return None
        
    return ruta_entrada

def analisis_exploratorio(df):
    """
    Realiza un análisis exploratorio básico usando series de numpy
    """
    # Convertir datos a arrays de numpy
    columnas_numericas = df.select_dtypes(include=[np.number]).columns
    
    # Crear series de numpy para cada aspecto del análisis
    series_stats = {}
    
    # 1. Estadísticas básicas por columna numérica
    for columna in columnas_numericas:
        datos = df[columna].values
        series_stats[columna] = np.array([
            np.mean(datos),      # Media
            np.median(datos),    # Mediana
            np.std(datos),       # Desviación estándar
            np.min(datos),       # Mínimo
            np.max(datos),       # Máximo
            np.percentile(datos, 25),  # Q1
            np.percentile(datos, 75)   # Q3
        ])
    
    # 2. Análisis de valores nulos
    nulos_series = {}
    for columna in df.columns:
        nulos = df[columna].isnull().sum()
        porcentaje = (nulos/len(df)) * 100
        nulos_series[columna] = np.array([nulos, porcentaje])
    
    # 3. Detección de outliers
    outliers_series = {}
    for columna in columnas_numericas:
        datos = df[columna].values
        Q1 = np.percentile(datos, 25)
        Q3 = np.percentile(datos, 75)
        IQR = Q3 - Q1
        limite_inferior = Q1 - 1.5 * IQR
        limite_superior = Q3 + 1.5 * IQR
        outliers = datos[(datos < limite_inferior) | (datos > limite_superior)]
        outliers_series[columna] = np.array([
            len(outliers),
            limite_inferior,
            limite_superior,
            np.mean(outliers) if len(outliers) > 0 else 0
        ])
    
    # Imprimir resultados usando las series
    print("\n=== ANÁLISIS ESTADÍSTICO POR COLUMNA ===")
    for columna, stats in series_stats.items():
        print(f"\nEstadísticas para {columna}:")
        print(f"Media: {stats[0]:.2f}")
        print(f"Mediana: {stats[1]:.2f}")
        print(f"Desv. Est.: {stats[2]:.2f}")
        print(f"Rango: [{stats[3]:.2f} - {stats[4]:.2f}]")
        print(f"IQR: [{stats[5]:.2f} - {stats[6]:.2f}]")
    
    print("\n=== ANÁLISIS DE VALORES NULOS ===")
    for columna, nulos in nulos_series.items():
        if nulos[0] > 0:
            print(f"{columna}: {int(nulos[0])} nulos ({nulos[1]:.2f}%)")
    
    print("\n=== ANÁLISIS DE OUTLIERS ===")
    for columna, outliers in outliers_series.items():
        if outliers[0] > 0:
            print(f"\n{columna}:")
            print(f"Cantidad de outliers: {int(outliers[0])}")
            print(f"Rango normal: [{outliers[1]:.2f} - {outliers[2]:.2f}]")
            print(f"Media de outliers: {outliers[3]:.2f}")
    
    # Limpiar datos
    df_limpio = df.copy()
    df_limpio = df_limpio.drop_duplicates()
    
    for columna in df_limpio.columns:
        if df_limpio[columna].isnull().sum() > 0:
            if df_limpio[columna].dtype in ['int64', 'float64']:
                df_limpio[columna] = df_limpio[columna].fillna(np.median(df_limpio[columna].values))
            else:
                df_limpio[columna] = df_limpio[columna].fillna('Sin especificar')
    
    # Retornar tanto el DataFrame limpio como las series de numpy
    return df_limpio, {
        'estadisticas': series_stats,
        'nulos': nulos_series,
        'outliers': outliers_series
    }

def exportar_resultados_excel(series_resultados, df_limpio, ruta_salida):
    """
    Exporta los resultados del análisis a un archivo Excel con formato
    """
    writer = pd.ExcelWriter(ruta_salida, engine='openpyxl')
    
    # Colores
    color_header = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    color_alerta = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    color_bueno = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # 1. Hoja de Estadísticas
    stats_rows = []
    for columna, stats in series_resultados['estadisticas'].items():
        stats_rows.append({
            'Columna': columna,
            'Media': stats[0],
            'Mediana': stats[1],
            'Desv. Est.': stats[2],
            'Mínimo': stats[3],
            'Máximo': stats[4],
            'Q1': stats[5],
            'Q3': stats[6]
        })
    
    df_stats = pd.DataFrame(stats_rows)
    df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
    
    # 2. Hoja de Valores Nulos
    nulos_rows = []
    for columna, nulos in series_resultados['nulos'].items():
        nulos_rows.append({
            'Columna': columna,
            'Cantidad Nulos': nulos[0],
            'Porcentaje': nulos[1]
        })
    
    df_nulos = pd.DataFrame(nulos_rows)
    df_nulos.to_excel(writer, sheet_name='Valores Nulos', index=False)
    
    # 3. Hoja de Outliers
    outliers_rows = []
    for columna, outliers in series_resultados['outliers'].items():
        if outliers[0] > 0:
            outliers_rows.append({
                'Columna': columna,
                'Cantidad Outliers': outliers[0],
                'Límite Inferior': outliers[1],
                'Límite Superior': outliers[2],
                'Media Outliers': outliers[3]
            })
    
    df_outliers = pd.DataFrame(outliers_rows)
    df_outliers.to_excel(writer, sheet_name='Outliers', index=False)
    
    # Aplicar formato
    workbook = writer.book
    
    for sheet_name in ['Estadísticas', 'Valores Nulos', 'Outliers']:
        worksheet = workbook[sheet_name]
        
        # Formato de encabezados
        for cell in worksheet[1]:
            cell.fill = color_header
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Aplicar colores condicionales
        if sheet_name == 'Valores Nulos':
            for row in worksheet.iter_rows(min_row=2):
                if float(row[2].value) > 5:  # Si hay más de 5% de nulos
                    for cell in row:
                        cell.fill = color_alerta
        
        elif sheet_name == 'Outliers':
            for row in worksheet.iter_rows(min_row=2):
                if float(row[1].value) > 10:  # Si hay más de 10 outliers
                    for cell in row:
                        cell.fill = color_alerta
    
    # Guardar el archivo
    writer.close()
    print(f"\nReporte exportado exitosamente a: {ruta_salida}")


def main():
    try:
        # Verificar rutas y obtener ruta de entrada
        ruta_entrada = verificar_rutas('campana_ventas_marzo_25.csv')
        if not ruta_entrada:
            return 
        
        print("Leyendo datos del archivo CSV...")
        df = leer_datos_campanas(ruta_entrada)
        
        print("Procesando datos...")
        df = procesar_datos(df)
        
        print("\nRealizando análisis exploratorio...")
        df_limpio, series_resultados = analisis_exploratorio(df)
        
        # Exportar resultados a Excel
        fecha_actual = datetime.now().strftime('%Y%m%d')
        ruta_excel = f'campanas/analisis_exploratorio_{fecha_actual}.xlsx'
        exportar_resultados_excel(series_resultados, df_limpio, ruta_excel)
        
        # Guardar series de resultados
        fecha_actual = datetime.now().strftime('%Y%m%d')
        np.savez(f'campanas/resultados_analisis_{fecha_actual}.npz',
                 estadisticas=series_resultados['estadisticas'],
                 nulos=series_resultados['nulos'],
                 outliers=series_resultados['outliers'])
        
        # Guardar datos limpios
        nombre_archivo = f'campanas_procesadas_limpias_{fecha_actual}.csv'
        ruta_salida = os.path.join('campanas', nombre_archivo)
        df_limpio.to_csv(ruta_salida, index=False, encoding='utf-8-sig')
        
        print(f"\nDatos limpios guardados en: {ruta_salida}")
        print(f"Series de resultados guardadas en: campanas/resultados_analisis_{fecha_actual}_*.npy")
        print(f"Total de campañas procesadas: {len(df)}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main() 